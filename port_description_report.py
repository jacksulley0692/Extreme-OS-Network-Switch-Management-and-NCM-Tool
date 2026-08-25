# -*- coding: utf-8 -*-
"""
Port description suggestion report - with uplink highlighting.

Based on port_description_report.py. Same read-only data gathering:
  - show ports description        (current descriptions, for comparison)
  - show lldp neighbors detail    (neighbor hostname, description, capabilities)
  - show edp neighbor             (Extreme-specific, catches Extreme-to-Extreme links)
  - show fdb                      (learned MACs per port, used as a last resort)

The one change: output is now an .xlsx (not .csv), and any port flagged as a
likely switch-to-switch uplink - i.e. multiple MACs seen on the port with no
LLDP/EDP neighbor identified - is highlighted in red. A plain .csv can't
carry cell colours, so highlighting requires a real spreadsheet file.

Nothing is written back to the switches - this only produces a report for
you to review. Writing suggested descriptions back is a deliberate next
phase, not part of this script.
"""

import logging
import configparser
import json
import re
from datetime import datetime
from pathlib import Path
from netmiko import ConnectHandler, NetmikoTimeoutException, NetmikoAuthenticationException
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

# =========================
# PATHS / CONFIG (reuses the same config.ini as switch_backup.py)
# =========================

SCRIPT_DIR = Path(__file__).resolve().parent

SWITCH_CANDIDATES = [
    SCRIPT_DIR / "Switches.txt",
    SCRIPT_DIR / "switches.txt",
    SCRIPT_DIR / "BackupScriptFiles" / "Switches.txt",
    SCRIPT_DIR / "BackupScriptFiles" / "switches.txt",
    SCRIPT_DIR.parent / "BackupScriptFiles" / "Switches.txt",
    SCRIPT_DIR.parent / "BackupScriptFiles" / "switches.txt",
    Path("/opt/switch-backup/BackupScriptFiles/Switches.txt"),
    Path("/opt/switch-backup/BackupScriptFiles/switches.txt"),
    Path("/opt/switch-backup/Switches.txt"),
    Path("/opt/switch-backup/switches.txt")
]

SWITCH_FILE = next((p for p in SWITCH_CANDIDATES if p.exists()), SCRIPT_DIR / "Switches.txt")

CONFIG_CANDIDATES = [
    SCRIPT_DIR / "conf.ini",
    SCRIPT_DIR / "config.ini",
    SCRIPT_DIR / "BackupScriptFiles" / "conf.ini",
    SCRIPT_DIR / "BackupScriptFiles" / "config.ini",
    SCRIPT_DIR.parent / "BackupScriptFiles" / "conf.ini",
    SCRIPT_DIR.parent / "BackupScriptFiles" / "config.ini",
    Path("/opt/switch-backup/BackupScriptFiles/conf.ini"),
    Path("/opt/switch-backup/BackupScriptFiles/config.ini"),
    Path("/opt/switch-backup/conf.ini"),
    Path("/opt/switch-backup/config.ini")
]

CONFIG_FILE = next((p for p in CONFIG_CANDIDATES if p.exists()), SCRIPT_DIR / "config.ini")

REPORT_FOLDER = SCRIPT_DIR / "reports"
LOG_FOLDER = SCRIPT_DIR / "logs"
STATUS_JSON = SCRIPT_DIR / "status.json"
STATUS_TXT = SCRIPT_DIR / "status.txt"

REPORT_FOLDER.mkdir(exist_ok=True)
LOG_FOLDER.mkdir(exist_ok=True)


def update_status(script_name, status, current_switch="", index=0, total=0, action="", counts=None, started_at=None):
    now_str = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    pct = round((index / total) * 100, 1) if total > 0 else 0
    progress_str = f"{index}/{total} ({pct}%)" if total > 0 else "0%"

    data = {
        "script": script_name,
        "status": status,
        "started_at": started_at or now_str,
        "updated_at": now_str,
        "progress": progress_str,
        "current_switch": current_switch,
        "latest_action": action,
        "counts": counts or {}
    }

    try:
        with open(STATUS_JSON, "w", encoding="utf-8") as f:
            json.dump(data, f, indent=2)

        txt_content = (
            "==================================================\n"
            f" Script:         {script_name}\n"
            f" Status:         {status}\n"
            f" Started At:     {data['started_at']}\n"
            f" Updated At:     {data['updated_at']}\n"
            f" Progress:       {progress_str}\n"
            f" Current Switch: {current_switch or 'N/A'}\n"
            f" Latest Action:  {action or 'N/A'}\n"
        )
        if counts:
            txt_content += " Counts:         " + ", ".join(f"{k}={v}" for k, v in counts.items()) + "\n"
        txt_content += "==================================================\n"

        with open(STATUS_TXT, "w", encoding="utf-8") as f:
            f.write(txt_content)
    except Exception as e:
        logging.debug(f"Failed to write status file: {e}")

config = configparser.ConfigParser()
if not CONFIG_FILE.exists():
    raise FileNotFoundError(f"Missing config file: {CONFIG_FILE}. Looked in: {[str(p) for p in CONFIG_CANDIDATES]}")
config.read(CONFIG_FILE)

def get_cfg(section, key, fallback=None):
    if config.has_section(section) and config.has_option(section, key):
        return config.get(section, key)
    for s in [section.lower(), section.capitalize(), section.upper(), "DEFAULT", "default", "settings", "Settings", "credentials", "backup", "connection"]:
        if config.has_section(s) and config.has_option(s, key):
            return config.get(s, key)
    if config.has_option("DEFAULT", key):
        return config.get("DEFAULT", key)
    return fallback

USERNAME = get_cfg("credentials", "username", fallback="admin")
PASSWORD = get_cfg("credentials", "password", fallback="")
METHOD = (get_cfg("connection", "method", fallback="telnet") or "telnet").lower()
try:
    TIMEOUT = int(get_cfg("connection", "timeout", fallback="20") or "20")
except Exception:
    TIMEOUT = 20

run_timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
log_file = LOG_FOLDER / f"port_description_report_{datetime.now().strftime('%Y-%m-%d')}.log"
xlsx_report = REPORT_FOLDER / f"port_description_suggestions_{run_timestamp}.xlsx"

# Row highlight styling for different uplink confidence levels
UPLINK_STYLES = {
    "DEFINITE UPLINK": {
        "fill": PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"),
        "font": Font(name="Arial", color="9C0006", bold=True),
    },
    "PROBABLE UPLINK": {
        "fill": PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"),
        "font": Font(name="Arial", color="9C6500", bold=True),
    },
    "POSSIBLE UPLINK": {
        "fill": PatternFill(start_color="E0F2FE", end_color="E0F2FE", fill_type="solid"),
        "font": Font(name="Arial", color="0369A1", bold=False),
    },
    "UNUSED": {
        "fill": None,
        "font": Font(name="Arial", color="94A3B8", italic=True),
    },
    "ENDPOINT": {
        "fill": None,
        "font": Font(name="Arial", color="000000"),
    },
}

logging.basicConfig(
    filename=log_file,
    level=logging.INFO,
    format="%(asctime)s - %(levelname)s - %(message)s"
)
console = logging.StreamHandler()
console.setLevel(logging.INFO)
console.setFormatter(logging.Formatter("%(message)s"))
logging.getLogger().addHandler(console)


# =========================
# BRAND KEYWORDS
# =========================
BRAND_KEYWORDS = [
    "meraki", "cisco", "avaya", "polycom", "poly", "aruba", "ubiquiti", "unifi",
    "extreme", "ruckus", "zebra", "xerox", "apple", "vmware", "yealink",
    "grandstream", "mitel", "dell", "hewlett", "hp ", "netgear", "fortinet",
    "sonicwall", "axis", "hikvision", "dahua", "lenovo", "microsoft", "samsung",
    "brother", "canon", "lexmark", "epson", "netapp", "synology", "qnap",
]

# =========================
# SMALL BUILT-IN OUI (VENDOR) TABLE - fallback when there's no LLDP/EDP data
# =========================
OUI_VENDORS = {
    "3C0754": "Apple", "F0F61C": "Apple", "AC87A3": "Apple", "0017F2": "Apple",
    "B827EB": "Raspberry Pi", "DCA632": "Raspberry Pi", "E45F01": "Raspberry Pi",
    "000C29": "VMware", "005056": "VMware",
    "0050B6": "HP", "3C4A92": "HP", "9457A5": "HP",
    "001B63": "Cisco", "0004E2": "Cisco",
    "F02F74": "Ubiquiti", "245A4C": "Ubiquiti", "DC9FDB": "Ubiquiti",
    "00E02B": "Extreme", "00043F": "Extreme",
    "00D0C9": "Intel", "3C970E": "Intel",
    "00D861": "Avaya", "00095B": "Avaya",
    "F81654": "Polycom", "0004F2": "Polycom",
    "00238A": "Aruba", "94B40F": "Aruba",
    "001DD8": "Ruckus",
    "3C5AB4": "Dell", "F8B156": "Dell",
    "9C8E99": "Zebra",
    "00609C": "Xerox",
}

# =========================
# LLDP CAPABILITY -> DEVICE TYPE
# =========================
CAPABILITY_TYPE_MAP = [
    (re.compile(r"telephone", re.IGNORECASE), "phone"),
    (re.compile(r"wlan|access point", re.IGNORECASE), "AP"),
    (re.compile(r"router", re.IGNORECASE), "router"),
    (re.compile(r"repeater", re.IGNORECASE), "repeater"),
    (re.compile(r"bridge", re.IGNORECASE), "switch"),
    (re.compile(r"station only", re.IGNORECASE), "endpoint"),
]

BRAND_TYPE_HINTS = {
    "axis": "camera", "hikvision": "camera", "dahua": "camera",
    "meraki": "AP",
    "avaya": "phone", "polycom": "phone", "poly": "phone",
    "yealink": "phone", "grandstream": "phone", "mitel": "phone",
    "xerox": "printer", "brother": "printer", "canon": "printer",
    "lexmark": "printer", "epson": "printer", "zebra": "printer",
    "netapp": "storage", "synology": "storage", "qnap": "storage",
    "vmware": "server (virtual)",
}

MAC_PATTERN = re.compile(r"^([0-9A-Fa-f]{2}[:\-\.]){5}[0-9A-Fa-f]{2}$")
IP_PATTERN = re.compile(r"^\d{1,3}(\.\d{1,3}){3}$")


def looks_like_useless_name(name):
    if not name:
        return True
    name = name.strip()
    if not name or name.lower() in ("none", "n/a", "unknown", "default"):
        return True
    if MAC_PATTERN.match(name):
        return True
    if IP_PATTERN.match(name):
        return True
    if re.match(r"^[0-9A-Fa-f]{12}$", name):
        return True
    return False


BRAND_DISPLAY_NAMES = {
    "meraki": "Meraki", "cisco": "Cisco", "avaya": "Avaya", "polycom": "Polycom",
    "poly": "Polycom", "aruba": "Aruba", "ubiquiti": "Ubiquiti", "unifi": "Ubiquiti",
    "extreme": "Extreme", "ruckus": "Ruckus", "zebra": "Zebra", "xerox": "Xerox",
    "apple": "Apple", "vmware": "VMware", "yealink": "Yealink",
    "grandstream": "Grandstream", "mitel": "Mitel", "dell": "Dell",
    "hewlett": "HP", "hp ": "HP", "netgear": "Netgear", "fortinet": "Fortinet",
    "sonicwall": "SonicWall", "axis": "Axis", "hikvision": "Hikvision",
    "dahua": "Dahua", "lenovo": "Lenovo", "microsoft": "Microsoft",
    "samsung": "Samsung", "brother": "Brother", "canon": "Canon",
    "lexmark": "Lexmark", "epson": "Epson", "netapp": "NetApp",
    "synology": "Synology", "qnap": "QNAP",
}


def find_brand(text):
    if not text:
        return None
    lowered = text.lower()
    for keyword in BRAND_KEYWORDS:
        if keyword.strip() in lowered:
            return BRAND_DISPLAY_NAMES.get(keyword, keyword.strip().title())
    return None


def find_device_type_from_capabilities(capabilities_text):
    if not capabilities_text:
        return None
    for pattern, device_type in CAPABILITY_TYPE_MAP:
        if pattern.search(capabilities_text):
            return device_type
    return None


def guess_vendor_from_mac(mac):
    oui = mac.upper().replace(":", "").replace("-", "").replace(".", "")[:6]
    return OUI_VENDORS.get(oui)


def get_device_type_netmiko():
    return "extreme_exos" if METHOD == "ssh" else "extreme_exos_telnet"


def read_switches(filename):
    try:
        with open(filename, "r") as file:
            return [
                line.strip() for line in file
                if line.strip() and not line.strip().startswith("#")
            ]
    except FileNotFoundError:
        logging.error(f"Switches.txt not found here: {filename}")
        return []


def draw_progress_bar(current, total, bar_length=25):
    percent = float(current) / total
    filled = int(round(percent * bar_length))
    bar = "#" * filled + "-" * (bar_length - filled)
    return f"[{bar}] {int(round(percent * 100))}% ({current}/{total})"


# =========================
# PARSERS
# =========================

def parse_port_descriptions(output):
    descriptions = {}
    for line in output.splitlines():
        line_clean = line.strip()
        if not line_clean or not re.match(r"^\d+(?::\d+)?\b", line_clean):
            continue
        # Split by columns (two or more spaces)
        parts = re.split(r"\s{2,}", line_clean)
        port = parts[0].strip()
        desc = ""
        if len(parts) > 1:
            candidate = parts[1].strip()
            # If candidate is not a link state (READY/UP, UP, DOWN, etc.)
            if not re.match(r"^(READY/UP|READY/DOWN|READY|UP|DOWN|ACTIVE|DISABLED|ENABLED|AUTO|\d+M|\d+G|FULL|HALF)$", candidate, re.IGNORECASE):
                desc = candidate
        if desc.lower() in ("enable", "disable", "----", "ready/up", "ready/down"):
            desc = ""
        descriptions[port] = desc
    return descriptions


def parse_lldp_neighbors_detail(output):
    neighbors = {}
    current_port = None
    current = {}

    def flush():
        if current_port and current:
            neighbors[current_port] = current.copy()

    for line in output.splitlines():
        port_match = re.search(r"Port\s+(\d+(?::\d+)?)\b", line, re.IGNORECASE)
        if port_match and ("detected" in line.lower() or "neighbor" in line.lower()):
            flush()
            current_port = port_match.group(1)
            current = {"name": "", "description": "", "capabilities": ""}
            continue

        if current_port is None:
            continue

        name_match = re.search(r"System\s*Name\s*:?\s*(.+)", line, re.IGNORECASE)
        if name_match:
            current["name"] = name_match.group(1).strip()
            continue

        desc_match = re.search(r"System\s*Description\s*:?\s*(.+)", line, re.IGNORECASE)
        if desc_match:
            current["description"] = desc_match.group(1).strip()
            continue

        cap_match = re.search(r"System\s*Capabilities\s*:?\s*(.+)", line, re.IGNORECASE)
        if cap_match:
            current["capabilities"] = cap_match.group(1).strip()
            continue

    flush()
    return neighbors


def parse_lldp_neighbors_summary(output):
    neighbors = {}
    for line in output.splitlines():
        match = re.match(r"^\s*(\d+(?::\d+)?)\s+(\S.*?)\s{2,}", line)
        if match:
            port, remote_name = match.group(1), match.group(2)
            neighbors[port] = {"name": remote_name, "description": "", "capabilities": ""}
    return neighbors


def parse_edp_neighbors(output):
    neighbors = {}
    for line in output.splitlines():
        match = re.match(r"^\s*(\d+(?::\d+)?)\s+(\S+)", line)
        if match:
            port, remote_name = match.group(1), match.group(2)
            neighbors[port] = remote_name
    return neighbors


def parse_fdb(output):
    port_macs = {}
    mac_pattern = re.compile(r"([0-9A-Fa-f]{2}[:\-]){5}[0-9A-Fa-f]{2}")
    for line in output.splitlines():
        mac_match = mac_pattern.search(line)
        if not mac_match:
            continue
        tokens = line.split()
        if not tokens:
            continue
        port_token = tokens[-1]
        if re.match(r"^\d+(:\d+)?$", port_token):
            port_macs.setdefault(port_token, []).append(mac_match.group(0))
    return port_macs


def classify_uplink_and_suggest(port, lldp, edp, fdb):
    macs = fdb.get(port, [])
    mac_count = len(macs)

    # 1. Check EDP (Extreme Discovery Protocol - Extreme Switch to Extreme Switch)
    if port in edp:
        remote_name = edp[port]
        if not looks_like_useless_name(remote_name):
            return {
                "suggestion": f"Uplink: {remote_name}",
                "source": "EDP Neighbor",
                "brand": "Extreme",
                "device_type": "switch",
                "uplink_status": "DEFINITE UPLINK",
                "uplink_reason": f"EDP inter-switch neighbor detected ({remote_name})"
            }

    # 2. Check LLDP
    if port in lldp:
        info = lldp[port]
        name = info.get("name", "")
        desc = info.get("description", "")
        caps = info.get("capabilities", "")
        brand = find_brand(desc) or find_brand(name)
        device_type = find_device_type_from_capabilities(caps)

        # Check if LLDP explicitly advertises switch/router capabilities or switch keywords in name/desc
        is_switch_cap = bool(re.search(r"bridge|router|repeater", caps, re.IGNORECASE))
        is_switch_keyword = bool(re.search(r"switch|stack|exos|catalyst|procurve|nexus|aruba", f"{name} {desc}", re.IGNORECASE))

        if is_switch_cap or is_switch_keyword:
            dev_label = device_type or "switch"
            host_label = name if not looks_like_useless_name(name) else (brand or "Switch")
            return {
                "suggestion": f"Uplink: {host_label}",
                "source": "LLDP Neighbor",
                "brand": brand or "Network",
                "device_type": dev_label,
                "uplink_status": "DEFINITE UPLINK",
                "uplink_reason": f"LLDP neighbor identified as {dev_label} ({host_label})"
            }

        # Otherwise LLDP is a connected endpoint (e.g., Phone, AP, Printer)
        if not device_type and brand:
            device_type = BRAND_TYPE_HINTS.get(brand.lower())

        sug = ""
        if brand and device_type:
            sug = f"{brand} {device_type}"
        elif brand:
            sug = f"{brand} device"
        elif device_type:
            sug = device_type.capitalize()
        elif not looks_like_useless_name(name):
            sug = name
        else:
            sug = "Connected Device"

        return {
            "suggestion": sug,
            "source": "LLDP Neighbor",
            "brand": brand or "",
            "device_type": device_type or "",
            "uplink_status": "ENDPOINT",
            "uplink_reason": f"LLDP endpoint detected ({sug})"
        }

    # 3. FDB MAC count fallback (No LLDP or EDP neighbor detected)
    if mac_count == 0:
        return {
            "suggestion": "Unused / no link",
            "source": "No data",
            "brand": "",
            "device_type": "",
            "uplink_status": "UNUSED",
            "uplink_reason": "No MAC addresses learned and no LLDP/EDP"
        }
    elif mac_count == 1:
        vendor = guess_vendor_from_mac(macs[0])
        sug = f"{vendor} device" if vendor else "Unknown device"
        return {
            "suggestion": sug,
            "source": "FDB (single MAC)",
            "brand": vendor or "",
            "device_type": "",
            "uplink_status": "ENDPOINT",
            "uplink_reason": f"Single MAC learned ({sug})"
        }
    elif mac_count == 2:
        return {
            "suggestion": "Possible Uplink / Passthrough (2 MACs)",
            "source": "FDB (2 MACs)",
            "brand": "",
            "device_type": "passthrough",
            "uplink_status": "POSSIBLE UPLINK",
            "uplink_reason": "2 MAC addresses learned (likely Phone with PC passthrough, mini-hub, or dual-nic device)"
        }
    else:  # mac_count > 2
        return {
            "suggestion": f"Probable Uplink ({mac_count} MACs)",
            "source": "FDB (multiple MACs)",
            "brand": "",
            "device_type": "unmanaged_switch/trunk",
            "uplink_status": "PROBABLE UPLINK",
            "uplink_reason": f"{mac_count} MAC addresses learned with no LLDP/EDP (unmanaged switch, hypervisor, or non-LLDP switch)"
        }


# =========================
# MAIN PER-SWITCH LOGIC
# =========================

def gather_switch_ports(ip):
    device = {
        "device_type": get_device_type_netmiko(),
        "host": ip,
        "username": USERNAME,
        "password": PASSWORD,
        "timeout": TIMEOUT,
    }

    rows = []

    try:
        logging.info(f"[{ip}] Connecting...")
        connection = ConnectHandler(**device)

        switch_info = connection.send_command("show switch")
        hostname_match = re.search(r"(?:Sysname|System name|Switch):\s*(\S+)", switch_info, re.IGNORECASE)
        switch_hostname = hostname_match.group(1) if hostname_match else ip

        desc_output = connection.send_command("show ports")

        lldp_detail_output = connection.send_command("show lldp neighbors detail")
        logging.info(f"[{ip}] RAW OUTPUT - show lldp neighbors detail:\n{lldp_detail_output}")
        lldp = parse_lldp_neighbors_detail(lldp_detail_output)

        if not lldp:
            lldp_summary_output = connection.send_command("show lldp neighbors")
            lldp = parse_lldp_neighbors_summary(lldp_summary_output)

        edp_output = connection.send_command("show edp neighbor")
        fdb_output = connection.send_command("show fdb")

        connection.disconnect()

        descriptions = parse_port_descriptions(desc_output)
        edp = parse_edp_neighbors(edp_output)
        fdb = parse_fdb(fdb_output)

        all_ports = set(descriptions) | set(lldp) | set(edp) | set(fdb)

        for port in sorted(all_ports, key=lambda p: [int(x) for x in p.split(":")]):
            res = classify_uplink_and_suggest(port, lldp, edp, fdb)
            rows.append({
                "ip": ip,
                "switch_hostname": switch_hostname,
                "port": port,
                "current_description": descriptions.get(port, ""),
                "suggested_description": res["suggestion"],
                "uplink_status": res["uplink_status"],
                "uplink_reason": res["uplink_reason"],
                "brand": res["brand"],
                "device_type": res["device_type"],
                "source": res["source"],
                "fdb_mac_count": len(fdb.get(port, [])),
            })

        logging.info(f"[{ip}] Gathered {len(rows)} port(s)")

    except NetmikoTimeoutException:
        logging.error(f"[{ip}] Netmiko timeout")
    except NetmikoAuthenticationException:
        logging.error(f"[{ip}] Authentication failed")
    except Exception as error:
        logging.error(f"[{ip}] Failed: {error}")

    return rows


def write_xlsx(all_rows):
    fieldnames = [
        "ip", "switch_hostname", "port",
        "current_description", "suggested_description",
        "uplink_status", "uplink_reason",
        "brand", "device_type", "source", "fdb_mac_count"
    ]

    wb = Workbook()
    ws = wb.active
    ws.title = "Port Descriptions"

    ws.append(fieldnames)
    for cell in ws[1]:
        cell.font = Font(name="Arial", bold=True)

    for row in all_rows:
        ws.append([row[field] for field in fieldnames])

    status_col_index = fieldnames.index("uplink_status") + 1  # openpyxl columns are 1-based
    for row_cells in ws.iter_rows(min_row=2):
        status_value = row_cells[status_col_index - 1].value
        style = UPLINK_STYLES.get(status_value, UPLINK_STYLES["ENDPOINT"])
        
        for cell in row_cells:
            if style["fill"]:
                cell.fill = style["fill"]
            if style["font"]:
                cell.font = style["font"]

    for col_cells in ws.columns:
        max_len = max(len(str(c.value)) for c in col_cells if c.value is not None)
        ws.column_dimensions[col_cells[0].column_letter].width = min(max_len + 4, 60)

    wb.save(xlsx_report)


def main():
    start_time = datetime.now()
    started_at_str = start_time.strftime("%Y-%m-%d %H:%M:%S")

    switches = read_switches(SWITCH_FILE)
    if not switches:
        logging.error("No switches found. Check Switches.txt.")
        update_status("port_description_report.py", "FAILED", action="No switches found in Switches.txt", started_at=started_at_str)
        return

    all_rows = []
    total = len(switches)

    update_status("port_description_report.py", "RUNNING", index=0, total=total, action="Starting port description report", started_at=started_at_str)

    logging.info("=" * 60)
    logging.info("Port description suggestion report (read-only)")
    logging.info(f"Switch file: {SWITCH_FILE}")
    logging.info(f"XLSX report: {xlsx_report}")
    logging.info("=" * 60)

    for index, ip in enumerate(switches, start=1):
        pbar = draw_progress_bar(index, total)
        logging.info("-" * 60)
        logging.info(f"PROGRESS {pbar} | Gathering ports for: {ip}")

        update_status(
            "port_description_report.py", "RUNNING", current_switch=ip, index=index, total=total,
            action=f"Gathering ports for {ip}", counts={"total_ports_collected": len(all_rows)}, started_at=started_at_str
        )

        all_rows.extend(gather_switch_ports(ip))

    write_xlsx(all_rows)

    definite_cnt = sum(1 for r in all_rows if r["uplink_status"] == "DEFINITE UPLINK")
    probable_cnt = sum(1 for r in all_rows if r["uplink_status"] == "PROBABLE UPLINK")
    possible_cnt = sum(1 for r in all_rows if r["uplink_status"] == "POSSIBLE UPLINK")

    final_counts = {
        "switches_processed": total,
        "total_port_rows": len(all_rows),
        "definite_uplinks": definite_cnt,
        "probable_uplinks": probable_cnt,
        "possible_uplinks": possible_cnt
    }

    update_status(
        "port_description_report.py", "COMPLETED", index=total, total=total,
        action="Report generated successfully", counts=final_counts, started_at=started_at_str
    )

    logging.info("=" * 60)
    logging.info(f"Done. {len(all_rows)} port row(s) written.")
    logging.info(f"Definite Switch Uplinks: {definite_cnt}")
    logging.info(f"Probable Uplinks (>2 MACs): {probable_cnt}")
    logging.info(f"Possible Uplinks / Passthroughs (2 MACs): {possible_cnt}")
    logging.info(f"XLSX report: {xlsx_report}")
    logging.info("=" * 60)


if __name__ == "__main__":
    main()
